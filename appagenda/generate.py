#!/usr/bin/env python3

"""
This script is used to generate the excel file for
the conference schedule on the Whova app. This script
uses the classes defined in the `orderfile.py` module
from the NAACL 2019 schedule repository that is integrated
as a submodule in this repository under the `agenda`
directory. For more details, run `generate.py --help`
and refer to the `README.md` file in this directory.

Note that this code is specific to the format that
Whova expects in its Excel template for the schedule.
For other app platforms such as GuideBook, this code
will need to be modified.

Author: Nitin Madnani
Date: May, 2019
"""


import argparse
import csv
import json
import logging
import sys

from pathlib import Path

from openpyxl import load_workbook

_THIS_DIR = Path(__file__).absolute().parent
AGENDA_SUBMODULE_DIR = _THIS_DIR.parent.joinpath('agenda', 'code')
sys.path.append(str(AGENDA_SUBMODULE_DIR))

from orderfile import Agenda, SessionGroup, Session, Item
from metadata import ScheduleMetadata
from utils import (classify_attendees,
                   get_tracks_for_session,
                   write_rows_in_sheet_at_cell)


class AppAgenda(Agenda):
    """
    Class encapsulating the agenda for the Whova app.
    Inherits from `orderfile.Agenda` and adds a
    `to_rows()` method to convert the parsed agenda
    into a list of rows that can then be written
    out into the Whova agenda template and
    imported into the app via the Whova EMS.
    """

    def __init__(self, *args):
        super(AppAgenda, self).__init__(*args)

    def to_rows(self,
                metadata,
                pdf_links=False,
                video_links=False,
                plenary_info={}):
        """
        Convert agenda to rows compatible
        with the Whova Agenda template.

        Parameters
        ----------
        metadata : ScheduleMetadata
            An instance of `ScheduleMetadata`
            containing the title, authors,
            abstracts, and anthology URLs for
            each item, if applicable.
        pdf_links : bool, optional
            Whether to generate the links to the
            anthology and other PDFs where appropriate.
            Defaults to `False`.
        video_links : bool, optional
            Whether to generate the links to
            the talk videos.
            Defaults to `False`.
        plenary_info : dict, optional
            Optional dictionary containing
            additional info for some plenary
            sessions. Defaults to an empty
            dictionary.

        Returns
        -------
        agenda_rows : list
            A list containing rows for the full schedule.
        """

        agenda_rows = []

        # iterate over the days in the agenda that should
        # already be in chronological order
        for day_index, day in enumerate(self.days):

            # now iterate over each day's contents
            for content in day.contents:

                # if it's a `SessionGroup`, cast it to the
                # `AppSessionGroup` class so that we can then
                # call its `to_rows()` method; we do this by
                # monkeypatching the `__class__` attribute which
                # is fine since we are just adding new behavior
                # (methods), not new attributes.
                if isinstance(content, SessionGroup):
                    content.__class__ = AppSessionGroup
                    session_group_rows = content.to_rows(day,
                                                         self.event,
                                                         metadata,
                                                         pdf_links=pdf_links,
                                                         video_links=video_links,
                                                         plenary_info=plenary_info)
                    agenda_rows.extend(session_group_rows)

                # if it's a `Session`, then cast it to `AppSession`
                # and call its `to_rows()` method and save that to
                # the agenda rows.
                elif isinstance(content, Session):
                    content.__class__ = AppSession
                    session_rows = content.to_rows(day,
                                                   self.event,
                                                   metadata,
                                                   pdf_links=pdf_links,
                                                   video_links=video_links,
                                                   plenary_info=plenary_info)
                    agenda_rows.extend(session_rows)

        # convert the list to a string and return
        return agenda_rows

    @classmethod
    def validate_rows(cls, rows):
        """
        A class method to validate that the given rows
        contain the fields that are required by the
        Whova template. The required fields are : date
        (index 0), start time(1), end time (2), and session
        title (4) are present for ALL rows; if there are
        some bad rows, save their indices so that we can
        show errors later after we save the sheet.

        Parameters
        ----------
        rows : list of lists
            A list of lists, each containing the
            string-valued fields for each row

        Returns
        -------
        invalid_rows : list of lists
            The subset of the given rows that
            do not contain the required fields.
        """
        invalid_rows = []
        for idx, row in enumerate(rows):
            required_fields = row[:3] + [row[4]]
            try:
                assert all([field != '' for field in required_fields])
            except AssertionError:
                invalid_rows.append(26 + idx)

        return invalid_rows


class AppSessionGroup(SessionGroup):
    """
    Class encapsulating a session group for the Whova app.
    Inherits from `orderfile.SessionGroup` and adds a
    `to_rows()` method to convert the parsed group
    into rows that can then be inserted into the Whova
    agenda template.
    """

    def __init__(self):
        super(AppSessionGroup, self).__init__()

    def to_rows(self,
                day,
                event,
                metadata,
                pdf_links=False,
                video_links=False,
                plenary_info={}):
        """
        Convert session group to a list of rows compatible
        with the Whova agenda template.

        Parameters
        ----------
        day : orderfile.Day
            The `Day` instance on which the session
            group is scheduled.
        metadata : ScheduleMetadata
            An instance of `ScheduleMetadata`
            containing the title, authors,
            abstracts, and anthology URLs for
            each item, if applicable.
        index : int
            An index to be used in the HTML tags
            for the box representing this session group.
        pdf_links : bool, optional
            Whether to generate the links to the
            anthology and other PDFs where appropriate.
            Defaults to `False`.
        video_links : bool, optional
            Whether to generate the links
            to the talk videos.
            Defaults to `False`.
        plenary_info : dict, optional
            Optional dictionary containing
            additional info for some plenary
            sessions. Defaults to an empty
            dictionary.

        Returns
        -------
        generated_rows : list
            A list containing the rows for the
            session group. The fields in each row are:
            date, start time, end time, tracks, session title,
            location, description, authors, and a string
            indicating whether it's a session ("Session")
            or a sub-session ("Sub") where a sub-session
            denotes presentation items in a session.
        """

        # initialize the result variable
        generated_rows = []

        # iterate over the sessions in the group which should
        # already be in chronological order
        for session in self.sessions:

            # cast `Session` to `AppSession` to enable
            # the call to `to_rows()`.
            session.__class__ = AppSession

            # the sessions in session groups may not
            # have a start and end time defined in the
            # order file, so we need to inherit those
            # here since sessions _are_ displayed with
            # start and end times in the app
            if not session.start and not session.end:
                session.start = self.start
                session.end = self.end

            # call the respective `to_rows()` for the session
            # and save the rows
            session_rows = session.to_rows(day,
                                           event,
                                           metadata,
                                           pdf_links=pdf_links,
                                           video_links=video_links,
                                           plenary_info=plenary_info)
            generated_rows.extend(session_rows)

        return generated_rows


class AppSession(Session):
    """
    Class encapsulating a session for the Whova app.
    Inherits from `orderfile.Session` and adds a
    `to_rows()` method to convert the parsed session
    into rows that can then be written out into the
    Whova agenda template.

    """
    def __init__(self):
        super(AppSession, self).__init__()

    def to_rows(self,
                day,
                event,
                metadata,
                pdf_links=False,
                video_links=False,
                plenary_info={}):
        """
        Convert session to a list of rows compatible
        with the Whova agenda template.

        Parameters
        ----------
        day : orderfile.Day
            The `Day` instance on which the session
            is scheduled.
        metadata : ScheduleMetadata
            An instance of `ScheduleMetadata`
            containing the title, authors,
            abstracts, and anthology URLs for
            each item, if applicable.
        index : int, optional
            An index to be used in some of the HTML tags.
        pdf_links : bool, optional
            Whether to generate the links to the
            anthology and other PDFs where appropriate.
            Defaults to `False`.
        video_links : bool, optional
            Whether to generate the links to
            the talk videos.
            Defaults to `False`.
        plenary_info : dict, optional
            Optional dictionary containing
            additional info for some plenary
            sessions. Defaults to an empty
            dictionary.

        Returns
        -------
        generated_rows : list
            A list containing the rows values for
            the session. The fields in each row are:
            date, start time, end time, tracks, session title,
            location, description, authors, and a string
            indicating whether it's a session ("Session")
            or a sub-session ("Sub") where a sub-session
            denotes presentation items in a session.
        """
        # initialize the result variable
        generated_rows = []

        # convert the given day to a date string
        date = day.datetime.strftime('%m/%d/%Y')

        # initialize description, and authors to be empty
        description = ''
        authors = ''

        # initialize title to be the session title
        title = self.title

        # get the tracks for this session
        tracks = get_tracks_for_session(self, event)

        # for tutorials, we may not have the start
        # and end times defined in the order file but we
        # need them for the app; if so just get them from
        # the first session item
        if self.type == 'tutorial':
            if not self.start and not self.end:
                self.start = self.items[0].start
                self.end = self.items[0].end

        # the best paper session in the main conference
        # or paper sessions in workshops (both not part
        # of session groups) may not have start
        # and end times defined in the order file but we
        # need them for the app; if so just get them from
        # the first and the last item in that session
        elif self.type in ['paper', 'best_paper']:
            if not self.start and not self.end:
                self.start = self.items[0].start
                self.end = self.items[-1].end

            # for these sessions, the description is just
            # the name of the session chair
            description = '<p>Chair: {}</p>'.format(self.chair) if self.chair else ''

            # include the ID in the title if available
            # but only for the main conference
            if event == 'main' and self.id_:
                title = '{}: {}'.format(self.id_, title)

            # TODO: add session livetweeters if available

        # we also want to include the ID in titles
        # for poster sessions in the main conference
        elif self.type == 'poster' and event == 'main' and self.id_:
            title = '{}: {}'.format(self.id_, title)

        # next use the extra plenary info provided, if appropriate
        elif self.type == 'plenary':
            self.abstract = ''
            self.person = ''
            self.person_url = ''
            self.pdf_url = ''
            self.video_url = ''
            for session_prefix in plenary_info:
                if self.title.startswith(session_prefix):
                    (self.abstract,
                     self.person,
                     self.person_affiliation,
                     self.person_url,
                     self.pdf_url,
                     self.video_url) = plenary_info[session_prefix]
                    break

            # initialize the desciption to be the abstract
            description = '<p>{}</p>'.format(self.abstract)
            authors = self.person

            # Add paper links and video links if we are asked to
            # and if we have the actual links to add
            if pdf_links and self.pdf_url:
                description += ' [<a href="{}">PDF</>]'.format(self.pdf_url)
            if video_links and self.video_url:
                description += ' [<a href="{}">VIDEO</a>]'.format(self.video_url)

        # this is always a 'Session'
        session_or_sub = 'Session'

        # all sessions get a row except for the tutorial
        # session since they are more like session groups
        # for those we only care about the items
        if self.type != 'tutorial':
            generated_rows.append([date,
                                   self.start,
                                   self.end,
                                   tracks,
                                   title,
                                   self.location,
                                   description,
                                   authors,
                                   '',
                                   session_or_sub])

        for item in self.items:

            # cast `Item` to `AppItem` to enable
            # the call to `to_rows()`.
            item.__class__ = AppItem

            # for tutorials and posters, we may not
            # have the start and end times but we need
            # to show these in the app, so get them
            # from the containing sessions
            if item.type in ['tutorial', 'poster']:
                item.start = self.start
                item.end = self.end

            # call `to_rows()` on each item and save
            # the resulting rows
            generated_rows.append(item.to_rows(day,
                                               event,
                                               metadata,
                                               pdf_links=pdf_links,
                                               video_links=video_links))

        return generated_rows


class AppItem(Item):
    """
    Class encapsulating a presentation item for
    the Whova app. Inherits from `orderfile.Item` and
    adds a `to_html()` method to convert the item
    into a row for the Whova app.

    """
    def __init__(self):
        super(AppItem, self).__init__()

    def to_rows(self,
                day,
                event,
                metadata,
                pdf_links=False,
                video_links=False):
        """
        Convert item to rows for the Whova app
        with the Whova Agenda template.

        Parameters
        ----------
        metadata : ScheduleMetadata
            An instance of `ScheduleMetadata`
            containing the title, authors,
            abstracts, and anthology URLs for
            each item, if applicable.
        pdf_links : bool, optional
            Whether to generate the links to the
            anthology and other PDFs where appropriate.
            Defaults to `False`.
        video_links : bool, optional
            Whether to generate the links to
            the talk videos.
            Defaults to `False`.

        Returns
        -------
        item_row : list of str
            A list of containing row values for the item.
            The fields in the row are: start time, end time,
            tracks, session title, location, description,
            authors, and a string indicating whether it's
            a session ("Session") or a sub-session ("Sub")
            where a sub-session denotes presentation items
            in a session.
        """

        # convert the given day to a date string
        date = day.datetime.strftime('%m/%d/%Y')

        # get the metadata for the item
        item_metadata = metadata.lookup(self.id_, event=event)
        self.title = item_metadata.title
        self.authors = '; '.join(item_metadata.authors)
        self.pdf_url = item_metadata.pdf_url
        self.video_url = item_metadata.video_url

        # set the description to be the abstract
        description = '<p>{}</p>'.format(item_metadata.abstract)

        # for the main conference, compute the track for the
        # item which we can get based on the ID suffix; if we
        # do not have a suffix, the track is simply "Main"
        # which refers to the main conference
        if event == 'main':
            if self.id_.endswith('-srw'):
                tracks = 'SRW'
            elif self.id_.endswith('-tacl'):
                tracks = 'TACL'
            elif self.id_.endswith('-demos'):
                tracks = 'Demos'
            elif self.id_.endswith('-industry'):
                tracks = 'Industry'
            elif self.id_.endswith('-tutorial'):
                tracks = 'Tutorial'
            else:
                tracks = 'Main'

        # for all items from workshops or co-located events
        # the only track is simply the name of the event
        else:
            tracks = event

        # Add paper links and video links if we are asked to
        # and if we have the actual links to add
        if pdf_links and self.pdf_url:
            description += ' [<a href="{}">PDF</>]'.format(self.pdf_url)
        if video_links and self.video_url:
            description += ' [<a href="{}">VIDEO</a>]'.format(self.video_url)

        # for everything except tutorials, we have "Sub"
        session_or_sub = 'Session' if self.type == 'tutorial' else 'Sub'

        # return the row
        return [date,
                self.start,
                self.end,
                tracks,
                self.title,
                self.location if hasattr(self, 'location') else '',
                description,
                self.authors,
                '',
                session_or_sub]


def main():

    # set up an argument parser
    parser = argparse.ArgumentParser(prog='generate.py')
    parser.add_argument("config_file",
                        help="Input JSON file containing "
                             "the app schedule configuration")
    parser.add_argument("output_agenda_file",
                        help="Output Excel file containing the app agenda")
    parser.add_argument("output_attendee_file",
                        help="Output Excel file containing non-speaker attendees")

    # parse given command line arguments
    args = parser.parse_args()

    # set up the logging
    logging.basicConfig(format='%(levelname)s - %(message)s', level=logging.INFO)

    # parse the configuration file
    with open(args.config_file, 'r') as configfh:
        config = json.loads(configfh.read())

    # parse the metadata files
    logging.info('Parsing metadata files ...')
    extra_metadata_files = config.get('extra_metadata_files', {})
    metadata = ScheduleMetadata.fromfiles(xmls=config['xml_files'],
                                          mappings=config['mapping_files'],
                                          extra_metadata_files=extra_metadata_files)

    # parse and store any additional plenary session
    # info if provided
    plenary_info_dict = {}
    if 'plenary_info_file' in config:
        logging.info("Parsing plenary info file ...")
        with open(config['plenary_info_file'], 'r') as plenaryfh:
            reader = csv.DictReader(plenaryfh, dialect=csv.excel_tab)
            for row in reader:
                key = row['session'].strip()
                value = (row['abstract'].strip(),
                         row['person'].strip(),
                         row['person_affiliation'].strip(),
                         row['person_url'].strip(),
                         row['pdf_url'].strip(),
                         row['video_url'].strip())
                plenary_info_dict[key] = value

    # parse the given order fiels into `AppAgenda` objects
    # and convert them to rows for the Whova app
    logging.info('Parsing order files and converting to rows...')
    agenda_rows = []
    for event, orderfile in config['order_files'].items():
        app_agenda = AppAgenda(event)
        app_agenda.fromfile(orderfile)
        rows = app_agenda.to_rows(metadata,
                                  pdf_links=config.get('pdf_links', False),
                                  video_links=config.get('video_links', False),
                                  plenary_info=plenary_info_dict)
        agenda_rows += rows

    # we need to manually add in the rows for the lunch and coffee
    # break sessions on June 6th and 7th; we removed them
    # from the workshop order files since it was leading to
    # Whova complain about duplicate rows
    agenda_rows.append(['06/06/2019', '7:30', '9:00', '',
                        'Breakfast', 'Hyatt Exhibit Hall & Nicollet Promenade',
                        '', '', '', 'Session'])
    agenda_rows.append(['06/06/2019', '10:30', '11:00', '',
                        'Morning coffee break', 'Hyatt Exhibit Hall & Nicollet Promenade',
                        '', '', '', 'Session'])
    agenda_rows.append(['06/06/2019', '12:30', '14:00', '',
                        'Lunch on your own', '',
                        '', '', '', 'Session'])
    agenda_rows.append(['06/06/2019', '15:30', '16:00', '',
                        'Afternoon coffee break', 'Hyatt Exhibit Hall & Nicollet Promenade',
                        '', '', '', 'Session'])
    agenda_rows.append(['06/07/2019', '7:30', '9:00', '',
                        'Breakfast', 'Hyatt Exhibit Hall & Nicollet Promenade',
                        '', '', '', 'Session'])
    agenda_rows.append(['06/07/2019', '10:30', '11:00', '',
                        'Morning coffee break', 'Hyatt Exhibit Hall & Nicollet Promenade',
                        '', '', '', 'Session'])
    agenda_rows.append(['06/07/2019', '12:30', '14:00', '',
                        'Lunch on your own', '',
                        '', '', '', 'Session'])
    agenda_rows.append(['06/07/2019', '15:30', '16:00', '',
                        'Afternoon coffee break', 'Hyatt Exhibit Hall & Nicollet Promenade',
                        '', '', '', 'Session'])

    # validate the rows and get the indices for the rows
    # that do not contain the required fields
    logging.info("Validating rows ...")
    invalid_rows = AppAgenda.validate_rows(agenda_rows)

    # at this point we need to hardcode some workarounds for
    # different people with the exact same names;

    # WORKAROUNDS:
    # (1) Paper "BERT Post-Training for Review Reading
    #     Comprehension and Aspect-based Sentiment
    #     Analysis": for this paper, "Bing Liu" from UIC
    #     is an author but there's an entirely different
    #     "Bing Liu" from Facebook AI who has registered
    #     for the conference but if we do not explicitly
    #     handle this case, Whova links the two so we need
    #     to somehow make them appear different; so we just
    #     add an asterisk to unregistered "Bing Liu" for now.
    # 
    # (2) Paper "A Soft Label Strategy for Target-Level Sentiment
    #     Classification": for this paper, "Xiao Liu" from Peking
    #     University is the author but another "Xiao Liu" from JHU
    #     is registered for the conference. Again, we just add
    #     an asterisk to the one from Peking University to
    #     distinguish between the two.
    for row in agenda_rows:
        title = row[4]
        if title == 'BERT Post-Training for Review Reading Comprehension and Aspect-based Sentiment Analysis':
            speaker_string = row[-3]
            new_speaker_string = speaker_string.replace('Bing Liu', "Bing Liu*")
            row[-3] = new_speaker_string
        elif title == 'A Soft Label Strategy for Target-Level Sentiment Classification':
            speaker_string = row[-3]
            new_speaker_string = speaker_string.replace('Xiao Liu', "Xiao Liu*")
            row[-3] = new_speaker_string

    # match the speakers in the agenda in the attendees
    # sheet and look up their metadata
    logging.info("Classifying attendees into speakers and non-speakers ...")
    df_speakers, df_non_speakers = classify_attendees(agenda_rows,
                                                      attendees_file=config.get('attendees_file', None))

    # read in the Whova agenda template
    logging.info('Populating Whova agenda template ...')
    workbook = load_workbook(str(_THIS_DIR / "Agenda_Track_Template.xlsx"))

    # write out the rows in "Agenda" sheet of the template
    sheet = workbook['Agenda']
    write_rows_in_sheet_at_cell(sheet, 'A26', agenda_rows)

    # now write out the speaker names and affiliations in the
    # "Speaker" sheet of the template
    sheet = workbook['Speaker']
    write_rows_in_sheet_at_cell(sheet, 'A6', df_speakers.to_numpy().tolist())

    # save the modified workbook to the given output file
    workbook.save(args.output_agenda_file)

    # show errors if have any missing required fields
    if len(invalid_rows) > 0:
        logging.error('The following rows in {} are missing '
                      'required fields: {}'.format(args.output_file,
                                                   invalid_rows))

    # now write out the non-speaker attendees file
    logging.info('Populating Whova attendee template with non-speakers ...')
    workbook = load_workbook(str(_THIS_DIR / "Attendee_list_template.xlsx"))
    sheet = workbook['Sheet1']
    attendee_rows = df_non_speakers[['Professional Name', 'Email', 'Affiliation']].to_numpy().tolist()
    attendee_rows = [['', ''] + row + ['', '', '', ''] for row in attendee_rows]
    write_rows_in_sheet_at_cell(sheet, 'A11', attendee_rows)
    workbook.save(args.output_attendee_file)


if __name__ == '__main__':
    main()
