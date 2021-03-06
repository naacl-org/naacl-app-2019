"""
This script contains utility functions needed to
support the main code in `generate.py` which is
used to generate the excel file fo the conference
schedule on the Whova app.

Author: Nitin Madnani
Date: May, 2019
"""

from openpyxl.utils import (get_column_letter,
                            range_boundaries)
from pandas import concat, DataFrame, read_excel

SESSION_TRACK_DISPLAY_DICT = {'srw': 'SRW',
                              'tacl': 'TACL',
                              'demos': 'Demos',
                              'industry': 'Industry'}


def write_rows_in_sheet_at_cell(sheet, cellstr, rows):
    """
    Write the given rows to the current sheet
    starting at the given cell. It is assumed
    that rows is is a list of lists with each
    child list being the same length.

    Note that nothing is returned since the `sheet`
    object is modified directly.

    Parameters
    ----------
    sheet :  openpyxl.worksheet.worksheet.Worksheet
        openpyxl Worksheet instance in which to
        write the given rows.
    cellstr : str
        Alphanumeric cell location in an Excel
        spreadsheet, e.g., 'A26'. The first element
        of the first list in `rows` is written at
        this location.
    rows : list of lists
        A list of lists with each list containing
        the same number of strings.
    """

    # figure out how many rows we are writing first
    num_rows_to_add = len(rows)

    # if there are any rows to add
    if num_rows_to_add > 0:

        # get the Cell object corresponding to the given
        # cell string
        starting_cell = sheet[cellstr]

        # compute the ending row index in the spreadsheet
        ending_row_index = starting_cell.row + num_rows_to_add - 1

        # compute the new column letter in the spreadsheet
        new_column_letter = get_column_letter(starting_cell.column + len(rows[0]) - 1)

        # get the cell range that we will be modifying
        cell_range = '{}:{}{}'.format(cellstr, new_column_letter, ending_row_index)

        # get the min and max rows and columns which openpyxl needs to
        # iterate over the rows
        (min_col, min_row, max_col, max_row) = range_boundaries(cell_range)

        # iterate over each spreadsheet row and write each data row
        # to the cells in the spreadsheet row
        for idx, sheet_row in enumerate(sheet.iter_rows(min_col=min_col,
                                                        max_col=max_col,
                                                        min_row=min_row,
                                                        max_row=max_row)):
            for cell, value in zip(sheet_row, rows[idx]):
                cell.value = value


def get_tracks_for_session(session, event):
    """
    Get the semicolon-separated tracks for a given
    session in a given event to be used in the Whova
    agenda template.

    Parameters
    ----------
    session : AppSession
        An `AppSession` object for which we want
        the tracks.
    event : str
        The event in which the session is being
        held.

    Returns
    -------
    tracks : str
        The semicolon-separated string indicating
        the tracks for this session. For workshops
        and co-located events, this is a single
        string with the same value as the event name.
        For the main conference – with the event value
        of 'main' – there can be multiple tracks.
    """
    # initialize the tracks to be empty string
    tracks = ''

    # for the main conference, if we have a paper session,
    # a poster session, or the best paper session, compute
    # the tracks based on the IDs; Note that "Research" is
    # always one of the tracks
    if event == 'main':
        if session.type in ['paper', 'poster', 'best_paper']:
            session_tracks = set()
            for item in session.items:
                try:
                    item_track = item.id_.split('-')[1]
                except IndexError:
                    session_tracks.add('Main')
                else:
                    session_tracks.add(SESSION_TRACK_DISPLAY_DICT[item_track])
                tracks = '; '.join(session_tracks)

    # for a workshop or a co-located event, we simply use
    # the name of the event as the only track name
    else:
        tracks = event

    return tracks


def classify_attendees(agenda_rows, attendees_file=None):
    """
    Given the rows containing agenda information, including
    speakers, and an optional file containing information
    about registered attendees, return two data frames:
    one containing information about speakers (whether or not
    they are registered) and the second containing information
    about registered attendees who are not speakers. The
    columns in both files are as follows: "Professional Name", "
    Email", and "Affiliation". However, note that for unregistered
    speakers, only the "Professional Name" field is populated
    as the other information is provided by the attendees file.

    Parameters
    ----------
    agenda_rows : list of lists
        A list of rows containing the fields for the App agenda.
    attendees_file : str, optional
        Path to the optional attendees file containing information
        for conference registrants.

    Returns
    -------
    (df_speakers, df_non_speaker_attendees) : tuple
        Tuple containing two data frames : the first
        containing information about speakers - whether
        registered or unregistered - and the second
        containing information about folk who are
        not speakers but are still registered for
        the conference.
    """
    # get all of the speakers in the agenda
    speakers = set()
    for row in agenda_rows:
        speaker_string = row[-3]
        if speaker_string:
            for speaker in speaker_string.split('; '):
                speakers.add(speaker)

    # if we are given an attendees file then
    # read that file into a data frame
    if attendees_file:
        df_attendees = read_excel(attendees_file,
                                  usecols=['Professional Name',
                                           'Affiliation',
                                           'Email'])

        # and then look up the emails and affiliations for the
        # speakers in the attendee info based on exact matches
        df_matched_speakers = df_attendees[df_attendees['Professional Name'].isin(speakers)]
        df_matched_speakers = df_matched_speakers[['Professional Name',
                                                   'Email',
                                                   'Affiliation']]
        # we also want the info for those attendees who are not speakers
        df_non_speaker_attendees = df_attendees[~df_attendees['Professional Name'].isin(speakers)].copy()

    # if no attendees file was provided, then we have no matching speakers
    # and no non-speaker attendees either.
    else:
        df_matched_speakers = DataFrame(columns=['Professional Name',
                                                 'Email',
                                                 'Affiliation'])
        df_non_speaker_attendees = DataFrame(columns=['Professional Name',
                                                      'Email',
                                                      'Affiliation'])

    # for those speakers who are not registered as attendees
    # just add their name and order the columns correctly
    missing_speaker_dicts = []
    for missing_speaker_name in speakers.difference(df_matched_speakers['Professional Name']):
        missing_speaker_dict = {'Professional Name': missing_speaker_name,
                                'Email': '',
                                'Affiliation': ''}
        missing_speaker_dicts.append(missing_speaker_dict)
    df_unmatched_speakers = DataFrame(missing_speaker_dicts)
    df_unmatched_speakers = df_unmatched_speakers[['Professional Name',
                                                   'Email',
                                                   'Affiliation']]

    # merge the matched and unmatched speaker data frames and drop the index
    df_speakers = concat([df_matched_speakers,
                          df_unmatched_speakers]).reset_index(drop=True)

    # drop any speakers that have the same name and email
    # since Whova does not like duplicates
    df_speakers.drop_duplicates(subset=['Professional Name', 'Email'],
                                inplace=True)

    # return the two data frames
    return df_speakers, df_non_speaker_attendees
